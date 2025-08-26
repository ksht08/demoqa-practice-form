import os
import pytest
from openpyxl import load_workbook

# pip install openpyxl

TMP_DIR = os.path.join(os.path.dirname(__file__), '..', 'TMP')
XLSX_FILE = "irregular verbs.xlsx"
XLSX_FILE_PATH = os.path.join(TMP_DIR, XLSX_FILE)
XLSX_SHEET_1 = "sheet 1"
XLSX_SHEET_2 = "Лист 2"

def test_read_xlsx_file():
    workbook = load_workbook(XLSX_FILE_PATH)

    sheet_names = workbook.sheetnames
    print("\n📄 Sheet names:")
    for sheet_name in sheet_names:
        print(f"{sheet_name}")

    sheet = workbook.active # get active sheet
    print(f"\n📄 Active sheet:\n{sheet.title}")
    if sheet.title == XLSX_SHEET_1:
        cell_c3 = sheet.cell(row=3, column=3).value
        print(f"\n𝄜 C3 value:\n{cell_c3}")
        assert cell_c3 == "awoken"
    elif sheet.title == XLSX_SHEET_2:
        cell_a17 = sheet.cell(row=17, column=1).value
        print(f"\n𝄜 A17 value:\n{cell_a17}")
        assert cell_a17 == "ячейка 17"
    
    
